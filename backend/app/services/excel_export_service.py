"""
Excel Export Service
Generates a 4-sheet Excel workbook from MonthlyOverviewResponse data.
Uses openpyxl with minimal but clean styling.
"""
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.monthly_overview import MonthlyOverviewResponse


# Sheet header colors (ARGB)
SHEET_COLORS = {
    "summary": "FF1E40AF",   # deep blue
    "debt": "FFB45309",      # amber
    "expense": "FFB91C1C",   # red
    "income": "FF15803D",    # green
}

HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def _make_header_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _auto_width(ws) -> None:
    """Auto-fit column widths based on cell content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _fmt_vnd(value: str) -> str:
    """Format a Decimal string as VND number."""
    try:
        return f"{int(Decimal(value)):,}"
    except Exception:
        return value


def generate_excel(data: MonthlyOverviewResponse, period_key: str) -> BytesIO:
    wb = Workbook()

    # ── Sheet 1: Tổng Quan ────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Tổng Quan"
    fill_s = _make_header_fill(SHEET_COLORS["summary"])

    headers_s = ["Chỉ Số", "Giá Trị (₫)"]
    for col_idx, h in enumerate(headers_s, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill_s
        cell.alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Tháng", period_key),
        ("Tổng Thu Nhập", _fmt_vnd(data.summary.total_income)),
        ("Tổng Chi Tiêu", _fmt_vnd(data.summary.total_expense)),
        ("Tổng Trả Nợ", _fmt_vnd(data.summary.total_debt_payment)),
        ("Dòng Tiền Ròng", _fmt_vnd(data.summary.net_cashflow)),
        ("Đã xử lý / Tổng", f"{data.summary.paid_count} / {data.summary.paid_count + data.summary.unpaid_count}"),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, 2):
        ws_summary.cell(row=row_idx, column=1, value=label)
        cell_v = ws_summary.cell(row=row_idx, column=2, value=value)
        cell_v.alignment = Alignment(horizontal="right")

    _auto_width(ws_summary)

    # ── Sheet 2: Nợ ──────────────────────────────────────────────────────────
    ws_debt = wb.create_sheet("Nợ")
    fill_d = _make_header_fill(SHEET_COLORS["debt"])
    debt_headers = ["Tên", "Loại", "Phân Loại", "Số Tiền Trả/Tháng (₫)", "Còn Lại (₫)", "Ngày Đến Hạn", "Trạng Thái"]
    for col_idx, h in enumerate(debt_headers, 1):
        cell = ws_debt.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill_d
        cell.alignment = Alignment(horizontal="center")

    debt_items = [i for i in data.items if i.source_type == "debt"]
    for row_idx, item in enumerate(debt_items, 2):
        ws_debt.cell(row=row_idx, column=1, value=item.name)
        ws_debt.cell(row=row_idx, column=2, value=item.category)
        # Loại column: distinguish personal loans from monthly installments
        loan_type_label = (
            "Vay cá nhân" if getattr(item, "debt_category", None) == "personal_lump_sum"
            else "Trả hàng tháng"
        )
        ws_debt.cell(row=row_idx, column=3, value=loan_type_label)
        c_amt = ws_debt.cell(row=row_idx, column=4, value=_fmt_vnd(item.amount))
        c_amt.alignment = Alignment(horizontal="right")
        c_rem = ws_debt.cell(row=row_idx, column=5, value=_fmt_vnd(item.remaining_amount or "0"))
        c_rem.alignment = Alignment(horizontal="right")
        ws_debt.cell(row=row_idx, column=6, value=item.due_day or "")
        ws_debt.cell(row=row_idx, column=7, value="✅ Đã Trả" if item.is_paid else "❌ Chưa Trả")


    _auto_width(ws_debt)

    # ── Sheet 3: Chi Tiêu ─────────────────────────────────────────────────────
    ws_expense = wb.create_sheet("Chi Tiêu")
    fill_e = _make_header_fill(SHEET_COLORS["expense"])
    expense_headers = ["Tên", "Loại", "Số Tiền (₫)", "Tần Suất", "Trạng Thái"]
    for col_idx, h in enumerate(expense_headers, 1):
        cell = ws_expense.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill_e
        cell.alignment = Alignment(horizontal="center")

    expense_items = [i for i in data.items if i.source_type == "expense"]
    for row_idx, item in enumerate(expense_items, 2):
        ws_expense.cell(row=row_idx, column=1, value=item.name)
        ws_expense.cell(row=row_idx, column=2, value=item.category)
        c_amt = ws_expense.cell(row=row_idx, column=3, value=_fmt_vnd(item.amount))
        c_amt.alignment = Alignment(horizontal="right")
        ws_expense.cell(row=row_idx, column=4, value=item.frequency)
        ws_expense.cell(row=row_idx, column=5, value="✅ Đã Chi" if item.is_paid else "❌ Chưa Chi")

    _auto_width(ws_expense)

    # ── Sheet 4: Thu Nhập ─────────────────────────────────────────────────────
    ws_income = wb.create_sheet("Thu Nhập")
    fill_i = _make_header_fill(SHEET_COLORS["income"])
    income_headers = ["Tên", "Loại", "Số Tiền (₫)", "Tần Suất"]
    for col_idx, h in enumerate(income_headers, 1):
        cell = ws_income.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill_i
        cell.alignment = Alignment(horizontal="center")

    income_items = [i for i in data.items if i.source_type == "income"]
    for row_idx, item in enumerate(income_items, 2):
        ws_income.cell(row=row_idx, column=1, value=item.name)
        ws_income.cell(row=row_idx, column=2, value=item.category)
        c_amt = ws_income.cell(row=row_idx, column=3, value=_fmt_vnd(item.amount))
        c_amt.alignment = Alignment(horizontal="right")
        ws_income.cell(row=row_idx, column=4, value=item.frequency)

    _auto_width(ws_income)

    # ── Save to BytesIO ───────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
